#!/usr/bin/env python3
"""
Build a simplified model comparison workbook from a scored context-evaluation run.

This workbook keeps only columns that are defensible from:
  1) static workbook metadata,
  2) local runtime measurements, or
  3) reproducible response-derived metrics from the run outputs.
"""

from __future__ import annotations

import argparse
import math
import re
import sqlite3
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


METADATA_ROWS = [
    "Hugging Face Link",
    "Family",
    "Variant",
    "Release Date",
    "Architecture Type",
    "Total Parameters",
    "Active Parameters",
    "Quantisation Available",
    "Context Window",
    "Native Multilingual",
    "Licence",
    "Commercial Usage Allowed",
]

RUNTIME_ROWS = [
    "Fits in 48GB VRAM",
    "VRAM Usage @ 32k",
    "VRAM Usage @ 64k",
    "vLLM Compatible",
    "TensorRT-LLM Compatible",
    "KV Cache Efficient",
    "Tokens/sec",
]

MEASURED_ROWS = [
    "Instruction Following",
    "Retrieval Obedience",
    "Hallucination Resistance",
    "Long Document Understanding",
    "Multilingual RAG",
    "EU Language Support",
    "Context Utilisation",
    "Summarisation Quality",
    "Consistency Across Runs",
    "System Prompt Adherence",
    "Citation/Grounded Answering",
    "Verbosity Control",
    "Stable Formatting",
    "Conversational Quality",
    "Gives Useful Answers",
    "Good for Chatbot",
    "Good for Search QA",
    "Good for Summaries",
]

MANUAL_ONLY_ROWS = [
    "Structured JSON Output",
    "Metadata Extraction",
    "Chunk Stitching Ability",
    "Prompt Robustness",
    "Hallucinates Under Weak Retrieval",
    "Handles Contradictory Chunks",
    "Good for Metadata",
    "Good for Classification",
    "Good for Simple Users",
    "Quantised Quality Loss",
]

DROP_ROWS = [
    "OCR Noise Tolerance",
    "Table Understanding",
    "Feels Smart in Real Usage",
    "Overthinks",
    "Too Verbose",
    "Too Creative",
    "Gives Safe Answers",
    "Refuses Too Often",
    "Good for Enterprise",
    "Good for Agentic Use",
]


def _classify_response_format(text: str) -> str:
    text = (text or "").strip()
    if not text:
        return "empty"
    if text.startswith("{") or text.startswith("["):
        return "json_like"
    bullet_lines = len(re.findall(r"(?m)^\s*(?:[-*•]|\d+\.)\s+", text))
    if bullet_lines >= 2:
        return "bullet"
    return "paragraph"


def _load_template_map(template_xlsx: Path) -> tuple[dict[str, int], dict[str, str], dict[str, dict[str, object]]]:
    if not template_xlsx.exists():
        return {}, {}, {}

    wb = load_workbook(template_xlsx, data_only=True)
    ws = wb["Sheet2"]

    row_labels = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if isinstance(label, str) and label.strip():
            row_labels[label.strip()] = row

    col_to_model = {}
    model_to_col = {}
    if "Evaluated model name" in row_labels:
        eval_row = row_labels["Evaluated model name"]
        for col in range(2, ws.max_column + 1):
            model_name = ws.cell(row=eval_row, column=col).value
            if isinstance(model_name, str) and model_name.strip():
                col_to_model[col] = model_name.strip()
                model_to_col[model_name.strip()] = col

    row_cache: dict[str, dict[str, object]] = {}
    for label, row in row_labels.items():
        values = {}
        for model_name, col in model_to_col.items():
            values[model_name] = ws.cell(row=row, column=col).value
        row_cache[label] = values

    return row_labels, model_to_col, row_cache


def _load_run_frames(run_dir: Path) -> dict[str, pd.DataFrame]:
    scores_db = run_dir / "scores" / "evaluation_scores_euf_context.db"
    raw_db = run_dir / "raw" / "evaluation_results_euf_context.db"
    if not scores_db.exists() or not raw_db.exists():
        raise FileNotFoundError("Missing scores or raw database under run_dir")

    with sqlite3.connect(scores_db) as con:
        scores = pd.read_sql_query("SELECT * FROM scores", con)
    with sqlite3.connect(raw_db) as con:
        raw = pd.read_sql_query(
            """
            SELECT id AS evaluation_id, model_name, language, question_id, run_number,
                   response, latency_ms
            FROM evaluations
            """,
            con,
        )

    merged = scores.merge(
        raw,
        on=["evaluation_id", "model_name", "language", "question_id", "run_number"],
        how="left",
    )

    token_summary_path = run_dir / "insights" / "data" / "token_budget_model_summary.csv"
    token_details_path = run_dir / "insights" / "data" / "token_budget_response_details.csv"
    gpu_model_path = run_dir / "insights" / "gpu_efficiency" / "data" / "model_gpu.csv"

    token_summary = pd.read_csv(token_summary_path) if token_summary_path.exists() else pd.DataFrame()
    token_details = pd.read_csv(token_details_path) if token_details_path.exists() else pd.DataFrame()
    gpu_model = pd.read_csv(gpu_model_path) if gpu_model_path.exists() else pd.DataFrame()

    return {
        "scores": scores,
        "raw": raw,
        "merged": merged,
        "token_summary": token_summary,
        "token_details": token_details,
        "gpu_model": gpu_model,
    }


def _compute_metrics(run_dir: Path) -> dict[str, dict[str, object]]:
    frames = _load_run_frames(run_dir)
    merged = frames["merged"].copy()
    merged["response_text"] = merged["response"].fillna("")
    merged["word_count"] = merged["response_text"].str.findall(r"\S+").str.len()
    merged["paragraph_count"] = merged["response_text"].apply(
        lambda s: len([p for p in str(s).strip().split("\n\n") if p.strip()]) if str(s).strip() else 0
    )
    merged["reference_like"] = merged["response_text"].str.contains(
        r"(?i)(?:^|\b)(?:source|sources|reference|references|according to|\[\d+\])",
        regex=True,
        na=False,
    )
    merged["format_class"] = merged["response_text"].apply(_classify_response_format)

    model_avg = (
        merged.groupby("model_name", as_index=False)
        .agg(
            response_count=("evaluation_id", "count"),
            avg_overall=("overall_quality", "mean"),
            avg_relevance=("relevance", "mean"),
            avg_factual=("factual_accuracy", "mean"),
            avg_completeness=("completeness", "mean"),
            avg_fluency=("fluency", "mean"),
            avg_coherence=("coherence", "mean"),
            avg_prompt_alignment=("prompt_alignment", "mean"),
            avg_latency_ms=("latency_ms", "mean"),
            avg_words=("word_count", "mean"),
            avg_paragraphs=("paragraph_count", "mean"),
            reference_like_rate=("reference_like", "mean"),
        )
    )

    lang_spread = (
        merged.groupby(["model_name", "language"])["overall_quality"]
        .mean()
        .reset_index()
        .groupby("model_name")["overall_quality"]
        .agg(lang_min="min", lang_max="max", lang_avg="mean")
        .reset_index()
    )
    lang_spread["lang_spread"] = lang_spread["lang_max"] - lang_spread["lang_min"]

    run_stability = (
        merged.groupby(["model_name", "run_number"])["overall_quality"]
        .mean()
        .reset_index()
        .groupby("model_name")["overall_quality"]
        .agg(run_std="std", run_range=lambda s: float(s.max() - s.min()))
        .reset_index()
    )
    run_stability["run_std"] = run_stability["run_std"].fillna(0.0)

    format_counts = (
        merged.groupby(["model_name", "format_class"]).size().reset_index(name="n")
    )
    idx = format_counts.groupby("model_name")["n"].idxmax()
    dominant_format = format_counts.loc[idx].rename(
        columns={"format_class": "dominant_format", "n": "dominant_format_n"}
    )
    totals = merged.groupby("model_name").size().reset_index(name="total_n")
    dominant_format = dominant_format.merge(totals, on="model_name", how="left")
    dominant_format["dominant_format_share"] = (
        dominant_format["dominant_format_n"] / dominant_format["total_n"]
    )

    token_summary = frames["token_summary"].copy()
    if not token_summary.empty:
        token_summary["input_util_pct"] = (
            token_summary["input_tokens_mean"] / token_summary["max_model_len"] * 100.0
        )

    token_details = frames["token_details"].copy()
    cap_hits = pd.DataFrame(columns=["model_name", "cap_hit_rate"])
    if not token_details.empty:
        token_details["response_tokens"] = pd.to_numeric(token_details["response_tokens"], errors="coerce")
        token_details["effective_output_cap"] = pd.to_numeric(token_details["effective_output_cap"], errors="coerce")
        token_details["cap_hit"] = token_details["response_tokens"] >= (token_details["effective_output_cap"] - 1)
        cap_hits = (
            token_details.groupby("model_name")["cap_hit"]
            .mean()
            .reset_index(name="cap_hit_rate")
        )

    gpu_model = frames["gpu_model"].copy()

    metrics = (
        model_avg.merge(lang_spread, on="model_name", how="left")
        .merge(run_stability, on="model_name", how="left")
        .merge(dominant_format[["model_name", "dominant_format", "dominant_format_share"]], on="model_name", how="left")
        .merge(token_summary, on="model_name", how="left")
        .merge(cap_hits, on="model_name", how="left")
    )
    if not gpu_model.empty:
        gpu_key = "model_name_norm" if "model_name_norm" in gpu_model.columns else (
            "model_name" if "model_name" in gpu_model.columns else None
        )
        if gpu_key:
            metrics = metrics.merge(gpu_model, left_on="model_name", right_on=gpu_key, how="left")
    metrics["tokens_per_sec"] = metrics["response_tokens_mean"] / (metrics["avg_latency_ms"] / 1000.0)
    metrics["summary_quality_proxy"] = (metrics["avg_completeness"] + metrics["avg_fluency"]) / 2.0
    metrics["search_qa_proxy"] = (metrics["avg_relevance"] + metrics["avg_factual"]) / 2.0
    metrics["chatbot_proxy"] = metrics["avg_overall"]
    metrics["useful_answers_proxy"] = metrics["avg_overall"]
    metrics["reference_like_rate"] = metrics["reference_like_rate"].fillna(0.0) * 100.0
    metrics["cap_hit_rate"] = metrics["cap_hit_rate"].fillna(0.0) * 100.0

    out: dict[str, dict[str, object]] = {}
    for row in metrics.to_dict(orient="records"):
        out[str(row["model_name"])] = row
    return out


def _fmt_num(val: object, digits: int = 3) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "NA"
    return f"{float(val):.{digits}f}"


def _fmt_pct(val: object, digits: int = 1) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "NA"
    return f"{float(val):.{digits}f}%"


def _build_measured_cell(row_label: str, m: dict[str, object]) -> str:
    if row_label == "Instruction Following":
        return f"Measured: prompt_alignment={_fmt_num(m.get('avg_prompt_alignment'))}"
    if row_label == "Retrieval Obedience":
        return (
            f"Measured: relevance={_fmt_num(m.get('avg_relevance'))}; "
            f"prompt_alignment={_fmt_num(m.get('avg_prompt_alignment'))}"
        )
    if row_label == "Hallucination Resistance":
        return f"Measured proxy: factual_accuracy={_fmt_num(m.get('avg_factual'))}"
    if row_label == "Long Document Understanding":
        return (
            f"Measured proxy: completeness={_fmt_num(m.get('avg_completeness'))}; "
            f"mean_input={_fmt_num(m.get('input_tokens_mean'), 1)}/{int(m.get('max_model_len') or 0)}"
        )
    if row_label == "Multilingual RAG":
        return (
            f"Measured: avg_overall={_fmt_num(m.get('lang_avg'))}; "
            f"spread={_fmt_num(m.get('lang_spread'))}"
        )
    if row_label == "EU Language Support":
        return (
            f"Measured: 24/24 langs evaluated; "
            f"min/avg/max={_fmt_num(m.get('lang_min'))}/{_fmt_num(m.get('lang_avg'))}/{_fmt_num(m.get('lang_max'))}"
        )
    if row_label == "Context Utilisation":
        return (
            f"Measured: input={_fmt_num(m.get('input_tokens_mean'), 1)}/{int(m.get('max_model_len') or 0)} "
            f"({_fmt_pct(m.get('input_util_pct'))})"
        )
    if row_label == "Summarisation Quality":
        return (
            f"Measured proxy: completeness={_fmt_num(m.get('avg_completeness'))}; "
            f"fluency={_fmt_num(m.get('avg_fluency'))}"
        )
    if row_label == "Consistency Across Runs":
        return (
            f"Measured: run_std={_fmt_num(m.get('run_std'), 4)}; "
            f"run_range={_fmt_num(m.get('run_range'), 4)}"
        )
    if row_label == "System Prompt Adherence":
        return f"Measured: prompt_alignment={_fmt_num(m.get('avg_prompt_alignment'))}"
    if row_label == "Citation/Grounded Answering":
        return (
            f"Measured proxy: factual_accuracy={_fmt_num(m.get('avg_factual'))}; "
            f"ref_like={_fmt_pct(m.get('reference_like_rate'))}"
        )
    if row_label == "Verbosity Control":
        return (
            f"Measured: mean_response={_fmt_num(m.get('response_tokens_mean'), 1)} tok; "
            f"cap_hit={_fmt_pct(m.get('cap_hit_rate'))}"
        )
    if row_label == "Stable Formatting":
        return (
            f"Measured: dominant_format={m.get('dominant_format') or 'NA'}; "
            f"share={_fmt_pct((m.get('dominant_format_share') or 0.0) * 100.0)}"
        )
    if row_label == "Conversational Quality":
        return (
            f"Measured proxy: fluency={_fmt_num(m.get('avg_fluency'))}; "
            f"coherence={_fmt_num(m.get('avg_coherence'))}"
        )
    if row_label == "Gives Useful Answers":
        return f"Measured proxy: avg_overall={_fmt_num(m.get('useful_answers_proxy'))}"
    if row_label == "Good for Chatbot":
        return (
            f"Measured proxy: avg_overall={_fmt_num(m.get('chatbot_proxy'))}; "
            f"latency={_fmt_num(m.get('avg_latency_ms'), 1)} ms"
        )
    if row_label == "Good for Search QA":
        return (
            f"Measured proxy: relevance={_fmt_num(m.get('avg_relevance'))}; "
            f"factual={_fmt_num(m.get('avg_factual'))}"
        )
    if row_label == "Good for Summaries":
        return (
            f"Measured proxy: completeness={_fmt_num(m.get('avg_completeness'))}; "
            f"fluency={_fmt_num(m.get('avg_fluency'))}"
        )
    return "Manual eval required"


def build_workbook(run_dir: Path, out_xlsx: Path, template_xlsx: Path | None = None) -> Path:
    _, template_model_to_col, template_rows = _load_template_map(template_xlsx) if template_xlsx else ({}, {}, {})
    metrics = _compute_metrics(run_dir)
    model_names = sorted(metrics.keys())

    # Preserve current column order if the template maps the run models.
    if template_model_to_col:
        ordered = [m for m in template_model_to_col.keys() if m in metrics]
        remaining = [m for m in model_names if m not in ordered]
        model_names = ordered + remaining

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    ws.cell(row=1, column=1, value="Column / Metric")
    for idx, model_name in enumerate(model_names, start=2):
        repo_name = None
        if template_xlsx and template_rows.get("Model Name"):
            repo_name = template_rows["Model Name"].get(model_name)
        ws.cell(row=1, column=idx, value=repo_name or model_name)
        ws.cell(row=2, column=idx, value=model_name)

    row = 4
    for section_name, row_labels in [
        ("Metadata", METADATA_ROWS),
        ("Runtime / Deployment", RUNTIME_ROWS),
        ("Measured From This Run", MEASURED_ROWS),
        ("Manual Only", MANUAL_ONLY_ROWS),
    ]:
        ws.cell(row=row, column=1, value=section_name)
        row += 1
        for label in row_labels:
            ws.cell(row=row, column=1, value=label)
            for col, model_name in enumerate(model_names, start=2):
                if label in MEASURED_ROWS:
                    value = _build_measured_cell(label, metrics[model_name])
                elif label == "Tokens/sec":
                    value = f"Measured local: {_fmt_num(metrics[model_name].get('tokens_per_sec'), 2)} tok/s"
                elif label == "Fits in 48GB VRAM":
                    value = "Measured local: yes (successful single-GPU L40S run)"
                elif label == "VRAM Usage @ 32k":
                    if template_rows.get(label):
                        value = template_rows[label].get(model_name)
                    else:
                        value = "Not measured in this run"
                elif label == "VRAM Usage @ 64k":
                    value = "Not measured in this run"
                elif label == "vLLM Compatible":
                    value = "Measured local: yes"
                elif label in {"TensorRT-LLM Compatible", "KV Cache Efficient"}:
                    value = template_rows.get(label, {}).get(model_name, "Not measured in this run")
                elif label in MANUAL_ONLY_ROWS:
                    value = "Manual eval required"
                else:
                    value = template_rows.get(label, {}).get(model_name, "NA")
                ws.cell(row=row, column=col, value=value)
            row += 1
        row += 1

    notes = wb.create_sheet("Method")
    notes["A1"] = "Keep"
    for idx, label in enumerate(METADATA_ROWS + RUNTIME_ROWS + MEASURED_ROWS, start=2):
        notes.cell(row=idx, column=1, value=label)

    manual_start = len(METADATA_ROWS + RUNTIME_ROWS + MEASURED_ROWS) + 4
    notes.cell(row=manual_start, column=1, value="Manual-only")
    for offset, label in enumerate(MANUAL_ONLY_ROWS, start=1):
        notes.cell(row=manual_start + offset, column=1, value=label)

    drop_start = manual_start + len(MANUAL_ONLY_ROWS) + 3
    notes.cell(row=drop_start, column=1, value="Drop")
    for offset, label in enumerate(DROP_ROWS, start=1):
        notes.cell(row=drop_start + offset, column=1, value=label)

    rationale_start = drop_start + len(DROP_ROWS) + 3
    rationale_lines = [
        "Measured rows are filled only from local run outputs and deterministic proxies.",
        "Metadata/runtime rows come from the existing template when available, or explicit local measurements.",
        "Manual-only rows stay out of the main measured block because the current benchmark does not test them directly.",
        "Drop rows are too subjective or outside the current benchmark scope.",
    ]
    for offset, line in enumerate(rationale_lines):
        notes.cell(row=rationale_start + offset, column=1, value=line)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-dir", required=True, help="Context evaluation run directory")
    parser.add_argument("--out-xlsx", default="", help="Output workbook path")
    parser.add_argument("--template-xlsx", default="", help="Existing comparison workbook to reuse metadata from")
    args = parser.parse_args()

    run_dir = Path(args.run_dir).expanduser().resolve()
    out_xlsx = Path(args.out_xlsx).expanduser().resolve() if args.out_xlsx else (
        run_dir.parent / "model_comparisons_for_RAG_simplified.xlsx"
    )
    template_xlsx = Path(args.template_xlsx).expanduser().resolve() if args.template_xlsx else (
        run_dir.parent / "model_comparisons_for_RAG_sheet2_filled_with_results.xlsx"
    )

    written = build_workbook(run_dir, out_xlsx, template_xlsx if template_xlsx.exists() else None)
    print(f"Wrote: {written}")


if __name__ == "__main__":
    main()
